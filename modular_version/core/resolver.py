import os
import re
import sqlite3
import time

from core.env import HAS_OPENPYXL
from core.db import init_db

if HAS_OPENPYXL:
    import openpyxl
else:
    openpyxl = None

def normalize_path_line(s: str) -> str:
    """Normalize path string by removing quotes and whitespace"""
    s = (s or "").strip()
    if not s:
        return ""
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        s = s[1:-1].strip()
    return s

def parse_paths_from_text(text: str) -> list[str]:
    """Parse text into list of unique paths"""
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = []
    for raw in text.split("\n"):
        raw = raw.strip()
        if not raw:
            continue
        if "\t" in raw:
            raw = raw.split("\t")[0].strip()
        p = normalize_path_line(raw)
        if p:
            lines.append(p)

    seen = set()
    out = []
    for p in lines:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out

def read_paths_from_txt(txt_path: str) -> list[str]:
    """Read paths from text file"""
    with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
        return parse_paths_from_text(f.read())

def read_paths_from_xlsx(xlsx_path: str) -> list[str]:
    """Read paths from Excel file"""
    if not HAS_OPENPYXL:
        raise RuntimeError("XLSX import requires openpyxl (pip install openpyxl).")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = [str(x).strip().lower() if x is not None else "" for x in row]
        break

    col_idx = 1
    if header_row:
        for i, h in enumerate(header_row, start=1):
            if h in ("path", "filepath", "file", "путь", "файл", "полный путь", "fullpath"):
                col_idx = i
                break

    paths = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) < col_idx:
            continue
        v = r[col_idx - 1]
        if v is None:
            continue
        p = normalize_path_line(str(v))
        if p:
            paths.append(p)

    wb.close()

    seen = set()
    out = []
    for p in paths:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


# ---------------------------
# resolve paths against DB
# ---------------------------

def resolve_one(con: sqlite3.Connection, raw_path: str) -> tuple[str, int | None, str | None, str | None]:
    """Resolve single path against database"""
    raw_path = normalize_path_line(raw_path)
    if not raw_path:
        return ("MISSING", None, None, "empty path")

    cur = con.cursor()

    row = cur.execute(
        "SELECT id, fullpath, is_present FROM files WHERE fullpath = ? LIMIT 1;",
        (raw_path,)
    ).fetchone()
    if row:
        return ("FOUND" if int(row[2]) == 1 else "MISSING", int(row[0]), str(row[1]), None)

    base = os.path.basename(raw_path)
    if not base:
        return ("MISSING", None, None, "no basename")

    hits = cur.execute(
        "SELECT id, fullpath FROM files WHERE name = ? AND is_present=1 LIMIT 500;",
        (base,)
    ).fetchall()

    if len(hits) == 1:
        return ("FOUND", int(hits[0][0]), str(hits[0][1]), "matched by name")
    if len(hits) > 1:
        return ("AMBIGUOUS", None, None, f"multiple matches by name: {len(hits)}")
    return ("MISSING", None, None, "not found")


def upsert_set(con: sqlite3.Connection, set_name: str) -> int:
    """Create or update a set"""
    now = int(time.time())
    cur = con.cursor()
    row = cur.execute("SELECT id FROM sets WHERE name=?;", (set_name,)).fetchone()
    if row:
        set_id = int(row[0])
        cur.execute("UPDATE sets SET updated_at=? WHERE id=?;", (now, set_id))
    else:
        cur.execute("INSERT INTO sets (name, created_at, updated_at) VALUES (?,?,?);", (set_name, now, now))
        set_id = int(cur.lastrowid)
    con.commit()
    return set_id


def sync_set_items(
    db_path: str,
    set_name: str,
    paths: list[str],
    *,
    replace: bool = True,
    status_cb=None,
    stop_flag=None
) -> dict:
    """Synchronize set items with database"""
    con = init_db(db_path, None)
    con.row_factory = sqlite3.Row
    set_id = upsert_set(con, set_name)
    cur = con.cursor()
    now = int(time.time())

    if replace:
        cur.execute("DELETE FROM set_items WHERE set_id=?;", (set_id,))
        con.commit()

    stats = {"total": 0, "found": 0, "missing": 0, "ambiguous": 0}

    batch = []
    batch_size = 300

    def flush():
        nonlocal batch
        if not batch:
            return
        cur.executemany(
            "INSERT INTO set_items (set_id, raw_path, file_id, resolved_path, status, note, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?);",
            batch
        )
        con.commit()
        batch = []

    for p in paths:
        if stop_flag and stop_flag.get("stop"):
            break

        status, file_id, resolved, note = resolve_one(con, p)
        if status == "FOUND":
            stats["found"] += 1
        elif status == "MISSING":
            stats["missing"] += 1
        else:
            stats["ambiguous"] += 1

        batch.append((set_id, p, file_id, resolved, status, note, now, now))
        stats["total"] += 1

        if len(batch) >= batch_size:
            flush()
            if status_cb:
                status_cb(
                    f"Set… {stats['total']} | FOUND {stats['found']} | MISS {stats['missing']} | AMB {stats['ambiguous']}"
                )

    flush()
    con.close()
    return stats
